import discord
import datetime
import asyncio
import urllib.request
import openpyxl
import random
import re
import youtube_dl


client = discord.Client()


@client.event
async def on_ready():
    print(client.user.id)
    print("Bot is ready")
    game = discord.Game("주인이온라인일때만작동")
    await client.change_presence(status=discord.Status.online, activity=game)


@client.event
async  def on_message(message):
    if message.content.startswith("비냥아"):
        co1 = "냐옹 먀옹 애옹 냥냥 먕먕 왜부르냥 왜"
        co1choice = co1.split(" ")
        co1number = random.randint(1, len(co1choice))
        co1result =co1choice[co1number-1]

        await message.channel.send(co1result)


    if message.content.startswith("비냥안녕"):
        co2 = "안녕하냥 안냥 반갑다냥 반갑냥 냐~옹 먀~옹"
        co2choice = co2.split(" ")
        co2number = random.randint(1, len(co2choice))
        co2result = co2choice[co2number -1]
        await message.channel.send(co2result)


    if message.content.startswith("비냥놀자"):
        co3 = "싫다냥 뭐하고놀거냥 가위바위보하자냥 혼자공놀이로충분하다냥 운동은질색이다냥"
        co3choice = co3.split(" ")
        co3number = random.randint(1, len(co3choice))
        co3result = co3choice[co3number -1]
        await message.channel.send(co3result)


    if message.content.startswith("비냥꺼져"):
        co4 = "ㅗ<^오^>ㅗ 너도꺼져라냥 흥!이다냥! 말걸지말라냥 집사호출한다냥?"
        co4choice = co4.split(" ")
        co4number = random.randint(1, len(co4choice))
        co4result = co4choice[co4number -1]
        await message.channel.send(co4result)


    if message.content.startswith("비냥울어"):
        co5 = "냐...아오옹... 먀..아오옹... 먀아아앙ㅠㅠ 냐아아앙ㅠㅠ 비냥이는...울지않는다냥.. 흑흑 집사냥ㅠㅠ"
        co5choice = co5.split(" ")
        co5number = random.randint(1, len(co5choice))
        co5result = co5choice[co5number -1]
        await message.channel.send(co5result)


    if message.content.startswith("비냥의모험"):
        await message.channel.send("등장인물 : 나무꾼, 비냥이, 일곱난쟁이, 왕냥이, 미냥이")


    if message.content.startswith("츄르"):
        co7 = "냥!!! 먕!!! 츄릅 츄르냥? 내놔라냥 츄르츄르! 내가제일좋아하는것이다냥! 고맙다냥 냥! 먕!"
        co7choice = co7.split(" ")
        co7number = random.randint(1, len(co7choice))
        co7result = co7choice[co7number -1]
        await message.channel.send(co7result)


    if message.content.startswith("비냥이손"):
        await message.channel.send("(촵)먀옹")

    if message.content.startswith("/캣타워인사"):
        await message.channel.send("```나의 캣타워다냐옹, 마음껏 사용해도 된다냥```")


    if message.content.startswith("비냥이귀여워"):
        co9 = "이놈의인기냥... 알고있다냥 (골골골골골~) 고맙다냥"
        co9choice = co9.split(" ")
        co9number = random.randint(1, len(co9choice))
        co9result = co9choice[co9number -1]
        await message.channel.send(co9result)


    if message.content.startswith("떼껄룩"):
        await message.channel.send("Take a look")


    if message.content.startswith("비냥키"):
        await message.channel.send("내 키는 17... 이다냥")

    if message.content.startswith("비냥번호좀"):
        co12 = "냥? 먕? 흠냥... 생각좀해보겠다냥 어디서내번호를!캬옹! 010 5664 9841 쉽게못알려주지냥 열심히쳐봐라냥 냥냥냐냥!"
        co12choice = co12.split(" ")
        co12number = random.randint(1, len(co12choice))
        co12result = co12choice[co12number -1]
        await message.channel.send(co12result)


    if message.content.startswith("/비냥게임"):
        await message.channel.send("```비연시를 시작하려면 /비연시시작'을 쳐주세요```")

    if message.content.startswith("/비연시시작"):
        await message.channel.send("```비냥이 연애 시뮬레이션을 시작합니다.```\n```/파트1 을 통해 스토리를 시작하세요!```\n```\n\n\n-정보-\n\n\n이름 : '비냥'\n\n나이 : 2\n\n특징 : 갈색정장에 흰색양말\n\n좋아하는것 : 와사비츄르, 생강츄르, 계피츄르 등```")



    if message.content.startswith("/파트1"):
        await message.channel.send("Part1 : 등굣길```\n\n\n\n\n\n나 : 우효~~ 내가 벌써 고등학생이라니.. 정말 기대되는걸?\n\n\n\n쿵!\n\n\n\n나 : 아얏!\n\n\n\n(반대편에서 오는 누군가와 부딪혔다)\n\n\n\n??? : 눈 똑바로 안뜨고 다니냥?!\n\n\n\n나 : 앗... 죄송합니다...\n\n\n\n??? : 쳇.. 옷만 더러워졌구냥... 오늘은 바쁘니 그냥간다냥...\n\n\n\n(먕 냥 댕 냥~)종소리가 울린다.\n\n\n\n나 : 으악! 지각이야 지각!!\n\n\n\n(헐레벌떡 가까스로 늦지않게 교실에 도착했다)\n\n\n\n나 : 휴.. 다행이도 안늦었다..  여기가 내가 지낼 교실이구나!  정말 기대되는걸?\n\n\n\n선생님 : 조용조용! 다들 자리에 앉아!\n\n\n\n나 : 저분이 우리반 담임선생님이구나... 정말 무섭다...\n\n\n\n선생님 : 출석 부르겠다... 음? 거기너 네 옆자리 누구야!\n\n\n\n나 : 저..저요?\n\n\n\n선생님 : 니 옆말고 빈자리가 있냐?\n\n\n\n나 : 저도 잘 모르겠습니다..\n\n\n\n그때 누군가가 문을 열고 들어온다.\n\n\n\n선생님 : 너 뭔데 지금 들어오냐 임마\n\n\n\n??? : 선생님이 알빠는 아니자냥...\n\n\n\n선생님 : 너 이 자식! 앞으로 나와!\n\n\n\n??? : 칫...\n\n\n\n(나 : 저 애는 아까 아침에 나랑 부딪힌 애다 기분이 매우 안좋아보이는데... 도와줄까?```\n```\n/도와준다\n\n또는\n\n/안도와준다\n\n를 입력하여 스토리를 진행하세요!```")


    if message.content.startswith("/안도와준다"):
        await message.channel.send("```(나 : 아침에 나한테 쌀쌀맞게 굴었으니 그냥 안도와줘야 겠다)\n\n\n\n선생님 : 너 아주 학교를 개판으로 다닐려고 하는구나? 너 이름이 뭐냐\n\n\n\n??? : 비냥이요..\n\n\n\n선생님 : 너 내가 화장실청소 1달 시키려했는데 이름이 귀여워서 2주만 시킨다. 들어가 앉아!\n\n\n\n비냥 : 칫...귀여운건 알아가지고...```")



    if message.content.startswith("/도와준다"):
        await message.channel.send("```나 : 선생님! 저 때문에 늦은거에요! 오늘 등교하는길에 저랑 부딪혀서 옷이 더러워졌거든요\n\n\n\n선생님 : 뭐? 거짓말 치지마!\n\n\n\n나 : 정말입니다 선생님 보세요 새옷이잖아요\n\n\n\n선생님 : 그러고 보니 새옷이구나... 크흠...  알았다 자리에 앉아\n\n\n\n??? : 누군지 모르지만 고맙다냥... 내이름은 비냥이다냥. 앞으로 편하게 부르도록하라냥\n\n\n\n```\n```비냥이와의 호감도 20상승 현재 호감도 20```\n```\n\n\n\n나 : 응응! 그래 비냥아!\n\n\n\n비냥 : 귀엽네냥...\n\n\n\n나 : 응? 뭐라했어?\n\n\n\n비냥 : 아...아무말도 안했다냥!\n\n\n\n그렇게 아침종례가 끝났다.\n\n\n\n나 : 음~ 학교나 한번 둘러볼까나~?\n\n\n\n??? : 거기... 귀여운 나의 아기 고양이여...\n\n\n\n토할거같다.\n\n\n\n나 : 저...저요?\n\n\n\n??? : 그래 너 말이야~ 여기에 너 말고 귀여운 아.기.고.양.이.가 어디있겠어~\n\n\n\n나 : 호에에..?\n\n\n\n??? : 초면에 실례를 범했군... 이 몸은 왕냥 이라고 하오...\n\n\n\n나 : 아.. 네! 왕냥님 안녕하세요..?\n\n\n\n 왕냥 : 오이오이... 그렇게 부담 안가져도 된다구~~\n\n\n\n나 : (엄청부담스럽다...)\n\n\n\n비냥 : 어이! 너!\n\n\n\n왕냥 : 음? 그대는 나를 부르는것인가?\n\n\n\n비냥 : 너 얘한테 뭐하는 짓이냥... 얘가 불편해 하자냥!\n\n\n\n왕냥 : 어이어이~ 왜 그리 화를 내는거야~? 난 그저 인사를 한 것 뿐이라구~\n\n\n\n비냥 : 뭐어어..! 인사??? 너 내 냥냥펀치 맛을 보고싶은거냥?!\n\n\n\n(나 : 나를 두고 비냥이와 왕냥이랑 싸우려고하네... 어떻게 말려야 하지...?)```\n\n```/1번 - 친구끼리 싸우지마아!```\n```/2번 - 왕냥아 너가 참아..```\n```/3번 비냥아 너가 참아..```")




    if message.content.startswith("비냥님"):
        food = "그래불렀느냥? 누가감히나를부르느냥 님'자를붙이다니..참으로착한인간이구냥 내가누군지알아보는구냥 내앞에고개를조아려라냥"
        foodchoice = food.split(" ")
        foodnumber = random.randint(1, len(foodchoice))
        foodresult = foodchoice[foodnumber-1]

        await message.channel.send(foodresult)

    if message.content.startswith("비냥탄신일"):
        foo = "고맙다냥! 감사하다냥! 냐냐냐냐냥! 먀먀먀먕!"
        foochoice = foo.split(" ")
        foonumber = random.randint(1, len(foochoice))
        fooresult = foochoice[foonumber-1]

        await message.channel.send(fooresult)




    if message.content.startswith("/비냥도움"):
        embed = discord.Embed(title="```비냥이 명령어```", description="```비냥이는 아직 미완성 입니다!```", color=0xff00ff)
        embed.add_field(name="```비냥아```", value="```그냥 비냥이를 부를때 사용하는 명령어```", inline=False)
        embed.add_field(name="```비냥놀자```", value="```비냥이를 놀아줘야 하지만 아직 준비중인 명령어```", inline=False)
        embed.add_field(name="```비냥안녕```", value="```비냥이한테 인사를 하는 명령어```", inline=False)
        embed.add_field(name="```비냥울어```", value="```비냥이를 울리는 명령어```", inline=False)
        embed.add_field(name="```비냥이손```", value="```손은 아니지만 손을 주는 명령어```", inline=False)
        embed.add_field(name="```비냥이귀여워```", value="```자동으로 골골송을 부르게 하는 명령어```", inline=False)
        embed.add_field(name="```비냥꺼져```", value="```흡사 티모를닮은 이모티콘으로 엿을 먹을수있는 명령어```", inline=False)
        embed.add_field(name="```비냥의모험```", value="```사실 이거 왜넣었는지 모르겠지만 추후 추가예정인 명령어```", inline=False)
        embed.add_field(name="```떼껄룩```", value="```Take a look```", inline=False)
        embed.add_field(name="```배고프다```", value="```비냥이에게 한마디 들을수있는 명령어```", inline=False)
        embed.add_field(name="```비냥님```", value="```비냥이의 집사를 넘어 노예가 될수있는 명령어```", inline=False)
        embed.add_field(name="```/가위,/바위,/보```", value="```비냥이와 가위바위보를 할수있는 명령어```", inline=False)
        embed.add_field(name="```/비냥게임```", value="```비냥이를 공략할수 있는 비연시를 할수있는 명령어```", inline=False)
        embed.add_field(name="```비냥봇은....```", value="```아직 미완성 입니다.\n앞으로 더 많은 기능이 추가될 예정입니다.\n숨겨진 명령어도 있으니 찾아보세요~```",
                        inline=False)


        embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/698723366185598978/700054070337077338/KakaoTalk_20200416_034435010.jpg")
        embed.set_footer(text="비냥봇 문의DM -비냥#3904-")
        await message.channel.send("요거슨 비냥이 명령어다냥!", embed=embed)




    if message.content.startswith("/정보"):
        date = datetime.datetime.utcfromtimestamp(((int(message.author.id) >> 22) + 1420070400000) / 1000)
        embed = discord.Embed(color=0xffff00)
        embed.add_field(name="이름", value=message.author.name, inline=True)
        embed.add_field(name="서버닉네임", value=message.author.display_name, inline=True)
        embed.add_field(name="가입일", value=str(date.year) + "년" + str(date.month) + "월" + str(date.day) + "일",
                                inline=False)
        embed.set_thumbnail(url=message.author.avatar_url)
        await message.channel.send(embed=embed)


    if message.content == "/가위" or message.content == "/바위" or message.content == "/보":
        bot_response = random.randint(1, 3)
        if bot_response == 1:
            if message.content == "/가위":
                await message.channel.send("가위! 앗 비겼다냥")
            elif message.content == "/바위":
                await message.channel.send("가위! 앗 내가졌다냥..")
            else:
                await message.channel.send("가위! 냥이쓰! 내가이겼다냥!")

        elif bot_response == 2:

            if message.content == "/가위":
                await message.channel.send("바위! 내가이겼다냥!")
            elif message.content == "/바위":
                await message.channel.send("바위! 비겼다냥")
            else:
                await message.channel.send("바위! 냐아앙...내가 졌다냥")
        else:
            if message.content == "/가위":
                await message.channel.send("보! 내가 졌다냥..")
            elif message.content == "/바위":
                await message.channel.send("보! 내가 이겼다냥!")
            else:
                await message.channel.send("보! 냥! 비겼다냥")



    if "시발" in message.content:
        file = openpyxl.load_workbook("경고.xlsx")
        sheet = file.active
        for i in range(1, 31):
            if str(sheet["A" + str(i)].value) == str(message.author.id):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                break
            if str(sheet["A" + str(i)].value) == "-":
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 1
                break
        file.save("경고.xlsx")

        await message.channel.send("너 비속어 경고다냥")

    if "병신" in message.content:
        file = openpyxl.load_workbook("경고.xlsx")
        sheet = file.active
        for i in range(1, 31):
            if str(sheet["A" + str(i)].value) == str(message.author.id):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                break
            if str(sheet["A" + str(i)].value) == "-":
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 1
                break
        file.save("경고.xlsx")

        await message.channel.send("너 비속어 경고다냥")

    if "닥쳐" in message.content:
        file = openpyxl.load_workbook("경고.xlsx")
        sheet = file.active
        for i in range(1, 31):
            if str(sheet["A" + str(i)].value) == str(message.author.id):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                break
            if str(sheet["A" + str(i)].value) == "-":
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 1
                break
        file.save("경고.xlsx")

        await message.channel.send("너 비속어 경고다냥")

    if "아가리" in message.content:
        file = openpyxl.load_workbook("경고.xlsx")
        sheet = file.active
        for i in range(1, 31):
            if str(sheet["A" + str(i)].value) == str(message.author.id):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                break
            if str(sheet["A" + str(i)].value) == "-":
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 1
                break
        file.save("경고.xlsx")

        await message.channel.send("너 비속어 경고다냥")
        
    if "집사바보" in message.content:
        file = openpyxl.load_workbook("경고.xlsx")
        sheet = file.active
        for i in range(1, 31):
            if str(sheet["A" + str(i)].value) == str(message.author.id):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                break
            if str(sheet["A" + str(i)].value) == "-":
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 1
                break
        file.save("경고.xlsx")

        await message.channel.send("감히 내 집사를 욕해?! 경고10회다냥")



    if message.content.startswith("/음악"):
        msg = message.content.split(" ")
        try:
            url = msg[1]
            url1 = re.match('(https?://)?(www\.)?((youtube\.(com))/watch\?v=([-\w]+)|youtu\.be/([-\w]+))', url)
            if url1 == None:
                await message.channel.send(message.channel, embed=discord.Embed(title=":no_entry_sign: url을 제대로 입력해주세요.",colour = 0x2EFEF7))
                return
        except IndexError:
            await message.channel.send(message.channel, embed=discord.Embed(title=":no_entry_sign: url을 입력해주세요.",colour = 0x2EFEF7))
            return

        channel = message.author.voice.voice_channel
        server = message.server
        voice_client = client.voice_clients_in(server)
        if client.is_voice_connected(server) and not playerlist[server.id].is_playing():
            await voice_client.disconnect()
        elif client.is_voice_connected(server) and playerlist[server.id].is_playing():
            player = await  voice_client.create_ytdl_player(url,after=lambda:queue(server.id),before_options="-reconnect 1 -reconnect_streamed 1 -reconnect_delay_max 5")
            if server.id in que:
                que[server.id].append(player)
            else:
                que[server.id] = [player]
            await client.send_message(message.channel, embed=discord.Embed(title=":white_check_mark: 추가 완료!",colour = 0x2EFEF7))
            platlist.append(player.title)
            return

        def queue(id):
            if que[id] !=[]:
                player = aue[id].pop(0)
                playerlist[id] = player
                del playlist[0]
                player.start()

        try:
            player = await voice_client.create_ytdl_player(url,after=lambda:queue(server.id),before_option="-reconnect 1 -reconnect_streamed 1 -reconnect_delay_max 5")
            playerlist[server.id] = player
            playlist.append(player.title)
        except youtube_dl.utils.DownloadError:
            await client.send_message(message.channel, embed=discord.Embed(title=":no_entry_sign: 존재하지 않는 경로입니다.",colour = 0x2EFEF7))
            await  voice_client.disconnect()
            return
        player.start()


    if message.content == "/종료":
        server = message.server
        voice_client = client.voice_client_in(server)

        if voice_client == None:
            await client.send_message(message.channel, embed=discord.Embed(title=":no_entry_sign: 봇이 음성채널에 없어요.",colour = 0x2EFEF7))
            return

        await client.send_message(message.channel, embed=discord.Embed(title=":mute: 채널에서 나갑니다.", colour = 0x2EFEF7))
        await voice_client.disconnect()

    if message.content == "/스킵":
        id = message.server.id
        if not playerlist[id].is_playing():
            await client.send_message(message.channel, embed=discord.Embed(title=":no_entry_sign: 스킵할 음악이 없어요.",colour = 0x2EFEF7))
            return
        await client.send_message(message.channel, embed=discord.Embed(title=":mute: 스킵했어요.",colour = 0x2EFEF7))
        playerlist[id].stop()

    if message.content == "/목록":

        if playlist == []:
            await client.send_message(message.channel, embed=discord.Embed(title=":no_entry_sign: 재생목록이 없습니다.",colour = 0x2EFEF7))
            return

        playstr = "```css\n[재생목록]\n\n"
        for i in range(0, len(playlist)):
            playstr += str(i+1)+" : "+playlist[i]+"\n"
        await client.send_message(message.channel, playstr + "```")






client.run("Njk4ODAyNTIwNTkxNjk1ODcy.XpL-iA.NGbrt4JcuVH6wL5lATr1cCID2jc")